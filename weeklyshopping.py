import csv
import datetime
import openpyxl
from openpyxl.utils.cell import get_column_letter
import pathlib


"""Dictionary Structure to Hold Data
+--------------+---------------+-------------------------------------------------------+
| <Sheet Name> | sheet_dates   | [date 1, date 2, ..., date n]                         |
|              +---------------+-------------------------------------------------------+
|              | sheet_headers | [header 1, header 2, ..., header n]                   |
|              +---------------+------------+------------+-----------------------------+
|              | data          |<Item ID>   | row        | <Row data from the file>    |
|              |               |            +------------+-----------------------------+
|              |               |            | pricedates | <date> | <unit price>       |
+--------------+---------------+------------+------------+--------+--------------------+
"""


def main():
    p = pathlib.Path('./data')
    sheets: dict[str, dict[str, list[str] |
                           str, dict[str, dict[str, list[str] |
                                               str, dict[str, float]]]]] = {}
    for file in [x for x in p.iterdir() if x.is_file()]:
        # print(file.stem, file.suffix)
        sheet_name: str = file.stem[0:-9].upper()
        sheet_date: str = file.stem[len(file.stem)-8:]
        # 'YYYY/MM/DD'
        sheet_date = sheet_date[4:] + "/" + sheet_date[0:2] + "/" + sheet_date[2:4]
        if sheet_name not in sheets.keys():
            sheets[sheet_name] = {'sheet_dates': [sheet_date],
                                  'data': {}}
        elif sheet_date not in sheets[sheet_name]['sheet_dates']:
            sheets[sheet_name]['sheet_dates'].append(sheet_date)

        data: list = []  # need to declare this outside the
        headers: list = []
        if file.suffix in ['.xls', '.xlsx']:
            wb = openpyxl.load_workbook(file)
            ws = wb.worksheets[0]
            data = list(ws.values)
            headers = list(data.pop(0))
            print(f'{sheet_name}\t{sheet_date}\t{file}\t{list(list(ws.values)[0])}')
        elif file.suffix in ['.csv', '.tsv']:
            print(f'{sheet_name}\t{sheet_date}\t{file}\t{list(list(ws.values)[0])}')
            with open(file, 'r', newline='', encoding='utf-8-sig') as csv_file:
                reader = csv.reader(csv_file)
                headers = list(next(reader))
                # print(row)
                data = list(reader)
        sheets[sheet_name]['sheet_headers'] = headers
        item_id = -1
        unit_price = -1
        if sheet_name == "FERGUSON":
            item_id = 3
            unit_price = 7
        elif sheet_name == "HOMEDEPOT":
            item_id = 7
            unit_price = 13
        elif sheet_name == "HOMEDEPOT-SECOND":
            item_id = 4
            unit_price = 10
        elif sheet_name == "HOMEDEPOTMABIS":
            item_id = 4
            unit_price = 10
        elif sheet_name == "LOWES":
            item_id = 4
            unit_price = 11
        else:
            print(f"I do not know about {sheet_name.title()}.")
            continue
        for row in data:
            item: str = row[item_id]
            if item not in sheets[sheet_name]['data'].keys():
                sheets[sheet_name]['data'][item] = {}
                sheets[sheet_name]['data'][item]['row'] = row
                sheets[sheet_name]['data'][item]['pricedates'] = {}
            try:
                if row[unit_price] is not None and isinstance(row[unit_price], str):
                    if row[unit_price].startswith('$'):
                        sheets[sheet_name]['data'][item]['pricedates'][sheet_date] = float(row[unit_price][1:].replace(',', ''))
                    elif row[unit_price].endswith('.¢'):
                        sheets[sheet_name]['data'][item]['pricedates'][sheet_date] = float('0.' + row[unit_price][:len(row[unit_price])-2].replace(',', ''))
                else:
                    if row[unit_price] is None or row[unit_price] == '':
                        sheets[sheet_name]['data'][item]['pricedates'][sheet_date] = '-'
                    else:
                        sheets[sheet_name]['data'][item]['pricedates'][sheet_date] = row[unit_price]
            except ValueError:
                sheets[sheet_name]['data'][item]['pricedates'][sheet_date] = row[unit_price]


    # print(d)
    wb = openpyxl.Workbook()
    temp_ws = wb.active

    for sheet, sheetdata in sheets.items():
        current_row = 1
        sheetdata['sheet_headers'].extend(sorted(sheetdata['sheet_dates']))
        sheetdata['sheet_headers'].insert(0, 'Changed Last Time')
        sheetdata['sheet_headers'].insert(0, 'Changed Sometime')
        ws = wb.create_sheet(sheet)
        ws.append(sheetdata['sheet_headers'])
        current_row += 1
        # for column, header in enumerate(sheetdata)
        # print(f'Sheetname : {sheet}\t', end='')
        for item, itemdata in sheetdata['data'].items():
            # print(f'Item : {item}\t', end='')
            row = list(itemdata['row'])
            for date in sheetdata['sheet_dates']:
                if date in sheets[sheet]['data'][item]["pricedates"].keys():
                    row.append(sheets[sheet]["data"][item]["pricedates"].get(date, "-"))
                else:
                    row.append('-')
                    # print(f'{sheets[sheet]["data"][item]["pricedates"].get(date,"")}', end='')
            # startdate = len(sheetdata['sheet_headers'])-len(sheetdata['sheet_dates'])
            # enddate = len(sheetdata['sheet_headers'])
            row.insert(0, '')
            #            + get_column_letter(enddate-1)
            #            + str(current_row)
            #            + '<>'
            #            + get_column_letter(enddate)
            #            + str(current_row))
            row.insert(0, '')
            #            + get_column_letter(startdate)
            #            + str(current_row)
            #            + ':'
            #            + get_column_letter(enddate)
            #            + str(current_row)
            #            + '},'
            #            + get_column_letter(enddate)
            #            + str(current_row)
            #            + '))'
            #            )
            #
            # # row.insert(0, '=average(n2:r2)=r2')
            ws.append(row)
            # ws.cell(row=current_row, column=0).formula
            current_row += 1
            if sheet == 'HOMEDEPOTMABIS':
                print(sheets[sheet]['data'][item]["pricedates"])
    wb.save('text.xlsx')


if __name__ == '__main__':
    main()

